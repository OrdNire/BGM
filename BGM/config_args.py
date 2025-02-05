import os
import numpy as np
import openpyxl
import math

def get_args_baseAR_infer(parser):
    # Meta
    parser.add_argument('--dataroot', type=str, default='../data/')
    parser.add_argument("--meta_root", type=str, default="../data/")
    parser.add_argument("--meta_file", type=str, default="SCSShip_retrieval.pkl")
    parser.add_argument("--code_file", type=str, default="autom_id_SCSShip_200.pkl")
    parser.add_argument('--dataset', type=str, default='SCSShip')
    parser.add_argument('--workers', type=int, default=16)
    parser.add_argument('--model_dir', type=str, default='/home/jiangfanzhuo/SpaceIR/baseline/results/baseAR_SCSShip_45_45_0.0001_0/')
    parser.add_argument('--model_name', type=str, default='ar_25.pth')

    # Noise
    parser.add_argument('--noise_rate', type=float, default=0)

    # Image Sizes
    parser.add_argument('--scale_size', type=int, default=256)
    parser.add_argument('--crop_size', type=int, default=224)

    # batch_size
    parser.add_argument('--batch_size', type=int, default=1)
    parser.add_argument('--test_batch_size', type=int, default=1)

    # inference
    parser.add_argument("--beam_size", type=int, default=100)

    # pipline name
    parser.add_argument('--name', type=str, default="baseAR_infer")

    args = parser.parse_args()

    args.model_path = os.path.join(args.model_dir, args.model_name)

    if args.dataset == 'DLRSD':
        args.num_labels = 17
    elif args.dataset == "MultiScene-Clean":
        args.num_labels = 36
    elif args.dataset == "FGSC-23":
        args.num_labels = 23
    elif args.dataset == "SCSShip":
        args.num_labels = 2
    else:
        print('dataset not included')
        exit()

    return args

def get_args_baseAR(parser):
    # Meta
    parser.add_argument('--dataroot', type=str, default='../data/')
    parser.add_argument("--meta_root", type=str, default="../data/")
    parser.add_argument("--meta_file", type=str, default="SCSShip_retrieval.pkl")
    parser.add_argument("--code_file", type=str, default="autom_id_SCSShip_200.pkl")
    parser.add_argument('--dataset', type=str, default='SCSShip')
    parser.add_argument('--workers', type=int, default=16)
    parser.add_argument('--results_dir', type=str, default='results/')
    parser.add_argument('--overwrite', action='store_true')
    parser.add_argument('--save_step', type=int, default=5)

    # Noise
    parser.add_argument('--noise_rate', type=float, default=0)

    # Optimization
    parser.add_argument('--batch_size', type=int, default=64)
    parser.add_argument('--test_batch_size', type=int, default=1)
    parser.add_argument('--indexing_epochs', type=int, default=45)
    parser.add_argument('--retrieval_epochs', type=int, default=45)
    parser.add_argument('--lr', type=float, default=1e-4)
    parser.add_argument('--smoothing', default=0.1, type=int, help='labelsmoothce smoothing')

    # Image Sizes
    parser.add_argument('--scale_size', type=int, default=256)
    parser.add_argument('--crop_size', type=int, default=224)

    # pipline name
    parser.add_argument('--name', type=str, default="baseAR")


    args = parser.parse_args()

    model_name = args.name + '_' + args.dataset

    model_name = model_name + "_" + str(args.indexing_epochs) + "_" + str(args.retrieval_epochs)

    model_name = model_name + '_' + str(args.lr)

    model_name = model_name + '_' + str(args.noise_rate)

    if args.dataset == 'DLRSD':
        args.num_labels = 17
    elif args.dataset == "MultiScene-Clean":
        args.num_labels = 36
    elif args.dataset == "FGSC-23":
        args.num_labels = 23
    elif args.dataset == "SCSShip":
        args.num_labels = 2
    else:
        print('dataset not included')
        exit()

    args.model_name = model_name
    args.save_path = os.path.join(args.results_dir, model_name)

    # if os.path.exists(args.save_path) and (not args.overwrite):
    #     print(args.save_path)
    #     overwrite_status = input('Already Exists. Overwrite?: ')
    #     if overwrite_status == 'rm':
    #         os.system('rm -rf ' + args.save_path)
    #     elif not 'y' in overwrite_status:
    #         exit(0)
    # elif not os.path.exists(args.save_path):
    os.makedirs(args.save_path, exist_ok=True)

    return args

def get_args_tokenizer(parser):
    parser.add_argument('--dataroot', type=str, default='./data/')
    parser.add_argument("--meta_root", type=str, default="./data/")
    parser.add_argument('--dataset', type=str, choices=["DLRSD"],
                        default='DLRSD')
    parser.add_argument('--workers', type=int, default=10)
    parser.add_argument('--results_dir', type=str, default='results/')
    parser.add_argument("--noise_rate", type=float, default=0.0)
    parser.add_argument('--test_known', type=int, default=0)

    # Optimization
    parser.add_argument('--optim', type=str, choices=['adam', 'sgd'], default='adam')
    parser.add_argument('--lr', type=float, default=0.0002)
    parser.add_argument('--batch_size', type=int, default=32)
    parser.add_argument('--test_batch_size', type=int, default=-1)
    parser.add_argument('--grad_ac_steps', type=int, default=1)
    parser.add_argument('--scheduler_step', type=int, default=1000)
    parser.add_argument('--scheduler_gamma', type=float, default=0.1)
    parser.add_argument('--epochs', type=int, default=20)
    parser.add_argument('--int_loss', type=float, default=0.0)
    parser.add_argument('--aux_loss', type=float, default=0.0)
    parser.add_argument('--loss_type', type=str, choices=['bce', 'mixed', 'class_ce', 'soft_margin'], default='bce')
    parser.add_argument('--scheduler_type', type=str, choices=['plateau', 'step'], default='plateau')
    parser.add_argument('--loss_labels', type=str, choices=['all', 'unk'], default='all')
    parser.add_argument('--lr_decay', type=float, default=0)
    parser.add_argument('--weight_decay', type=float, default=1e-4)
    parser.add_argument('--max_samples', type=int, default=-1)
    parser.add_argument('--max_batches', type=int, default=-1)
    parser.add_argument('--warmup_scheduler', action='store_true', help='')
    parser.add_argument('--margin', type=float, default=0.5)
    parser.add_argument('--beta', type=float, default=0.5)
    parser.add_argument('--contr_lambda', type=float, default=0.5)

    # Model
    parser.add_argument('--layers', type=int, default=3)
    parser.add_argument('--heads', type=int, default=4)
    parser.add_argument('--dropout', type=float, default=0.1)
    parser.add_argument('--pos_emb', action='store_true', help='positional encoding')
    parser.add_argument('--use_lmt', dest='use_lmt', action='store_true', help='label mask training')
    parser.add_argument('--freeze_backbone', action='store_true')
    parser.add_argument('--no_x_features', action='store_true')

    # Image Sizes
    parser.add_argument('--scale_size', type=int, default=640)
    parser.add_argument('--crop_size', type=int, default=576)

    # Testing Models
    parser.add_argument('--inference', action='store_true')
    parser.add_argument('--resume', action='store_true')
    parser.add_argument('--saved_model_name', type=str, default='')

    parser.add_argument('--overwrite', action='store_true')
    parser.add_argument('--name', type=str, default='')
    args = parser.parse_args()

    model_name = args.dataset
    if args.dataset == 'DLRSD':
        args.num_labels = 17
    else:
        print('dataset not included')
        exit()

    model_name += '.' + str(args.layers) + 'layer'
    model_name += '.bsz_{}'.format(int(args.batch_size * args.grad_ac_steps))
    model_name += '.' + args.optim + str(args.lr)  # .split('.')[1]

    if args.use_lmt:
        model_name += '.lmt'
        args.loss_labels = 'unk'
        model_name += '.unk_loss'
        args.train_known_labels = 100
    else:
        args.train_known_labels = 0

    if args.pos_emb:
        model_name += '.pos_emb'

    if args.int_loss != 0.0:
        model_name += '.int_loss' + str(args.int_loss).split('.')[1]

    if args.aux_loss != 0.0:
        model_name += '.aux_loss' + str(args.aux_loss).replace('.', '')

    if args.no_x_features:
        model_name += '.no_x_features'

    args.test_known_labels = int(args.test_known * 0.01 * args.num_labels)

    if args.name != '':
        model_name += '.' + args.name

    if not os.path.exists(args.results_dir):
        os.makedirs(args.results_dir)

    model_name = os.path.join(args.results_dir, model_name)

    args.model_name = model_name

    if args.inference:
        args.epochs = 1

    if os.path.exists(args.model_name) and (not args.overwrite) and (not 'test' in args.name) and (not eval) and (
    not args.inference) and (not args.resume):
        print(args.model_name)
        overwrite_status = input('Already Exists. Overwrite?: ')
        if overwrite_status == 'rm':
            os.system('rm -rf ' + args.model_name)
        elif not 'y' in overwrite_status:
            exit(0)
    elif not os.path.exists(args.model_name):
        os.makedirs(args.model_name)

    return args

def get_args_train_SIR(parser):
    parser.add_argument('--dataroot', type=str, default='./data/')
    parser.add_argument("--meta_root", type=str, default="./data/")
    parser.add_argument("--meta_file", type=str, default="DLRSD_retrieval_noise0.pkl")
    parser.add_argument('--dataset', type=str, choices=["DLRSD", "MultiScene-Clean", "ML-AID", "BigEarthNet-19", "WHDLD", "PlanetUAS"],
                        default='DLRSD')
    parser.add_argument('--workers', type=int, default=16)
    parser.add_argument('--results_dir', type=str, default='results/')
    parser.add_argument("--noise_rate", type=float, default=0.0)
    parser.add_argument("--cuda", type=str, default="0,1")

    # Optimization
    parser.add_argument('--optim', type=str, choices=['adam', 'sgd'], default='adam')
    parser.add_argument('--lr', type=float, default=3.5e-5)
    parser.add_argument("--batch_size", type=int, default=64)
    parser.add_argument("--classification_epochs", type=int, default=100)
    parser.add_argument("--indexing_epochs", type=int, default=50)
    parser.add_argument("--retrieval_epochs", type=int, default=50)
    parser.add_argument("--smoothing", type=float, default=0.1)
    parser.add_argument("--contr_lambda", type=float, default=0.01)
    # parser.add_argument("--contr_mu", type=float, default=0.1)
    parser.add_argument("--grad_ac_steps", type=int, default=1)
    parser.add_argument("--classification_resume", action="store_true")
    parser.add_argument("--classification_path", type=str, default="/home/jiangfanzhuo/SpaceIR/results/DLRSD.6layer.12heads.bsz_64.cls_eps_100.idx_eps_50.ret_eps_50.id_len_1.contr_l_0.1.contr_mu_0.1.distance_prc_10.class_loss_CE.adam3.5e-05.exp4_MLPdecoder_w_global/classification_model.pth")
    parser.add_argument("--indexing_resume", action="store_true")
    parser.add_argument("--indexing_path", type=str,
                        default="/home/jiangfanzhuo/SpaceIR/results/DLRSD.6layer.12heads.bsz_64.cls_eps_100.idx_eps_50.ret_eps_50.id_len_1.contr_l_0.1.contr_mu_0.1.distance_prc_10.class_loss_CE.adam3.5e-05.exp4_MLPdecoder_w_global/indexing_model.pth")
    parser.add_argument("--class_loss", type=str, default="CE")
    parser.add_argument("--test_batch_size", type=int, default=128)

    # Model
    parser.add_argument('--layers', type=int, default=6)
    parser.add_argument('--heads', type=int, default=12)
    parser.add_argument('--dropout', type=float, default=0.1)
    parser.add_argument("--distance_prc", type=int, default=5)
    parser.add_argument('--id_len', type=int, default=1)
    parser.add_argument('--k_bit', type=int, default=8)
    parser.add_argument('--hidden_dim', type=int, default=768)
    parser.add_argument("--decoder_type", type=str, default="MLP")

    # Image Sizes
    parser.add_argument('--scale_size', type=int, default=640)
    parser.add_argument('--crop_size', type=int, default=576)

    # Testing Models
    parser.add_argument('--inference', action='store_true')
    parser.add_argument('--resume', action='store_true')
    parser.add_argument('--saved_model_name', type=str, default='')

    parser.add_argument('--overwrite', action='store_true')
    parser.add_argument('--name', type=str, default='exp4_MLPdecoder_w_global')
    args = parser.parse_args()

    model_name = args.dataset
    if args.dataset == 'DLRSD':
        args.num_labels = 17
    elif args.dataset == "MultiScene-Clean":
        args.num_labels = 36
    elif args.dataset == "ML-AID":
        args.num_labels = 17
    elif args.dataset == "BigEarthNet-19":
        args.num_labels = 15
    elif args.dataset == "WHDLD":
        args.num_labels = 6
    elif args.dataset == "PlanetUAS":
        args.num_labels = 14
    else:
        print('dataset not included')
        exit()

    model_name += '.' + str(args.layers) + 'layer'
    model_name += '.' + str(args.heads) + 'heads'
    model_name += ".bsz_{}".format(int(args.batch_size))
    model_name += '.cls_eps_{}'.format(int(args.classification_epochs * args.grad_ac_steps))
    model_name += '.idx_eps_{}'.format(int(args.indexing_epochs * args.grad_ac_steps))
    model_name += '.ret_eps_{}'.format(int(args.retrieval_epochs * args.grad_ac_steps))
    model_name += ".id_len_{}".format(int(args.id_len))
    model_name += ".contr_l_{}".format(str(args.contr_lambda))
    # model_name += ".contr_mu_{}".format(str(args.contr_mu))
    model_name += ".distance_prc_{}".format(int(args.distance_prc))
    model_name += ".class_loss_{}".format(str(args.class_loss))
    model_name += '.' + args.optim + str(args.lr)  # .split('.')[1]

    if args.name != '':
        model_name += '.' + args.name

    if not os.path.exists(args.results_dir):
        os.makedirs(args.results_dir)

    model_name = os.path.join(args.results_dir, model_name)

    args.model_name = model_name

    if args.inference:
        args.epochs = 1

    # global feature(hash)
    # find the value of ζ

    # workbook = openpyxl.load_workbook('codetable.xlsx')
    # sheet = workbook.active
    # args.threshold = sheet.cell(row=64 + 1, column=math.ceil(math.log(args.num_labels, 2)) + 1).value
    args.threshold = 0.5

    if os.path.exists(args.model_name) and (not args.overwrite) and (not 'test' in args.name) and (not eval) and (
    not args.inference) and (not args.resume):
        print(args.model_name)
        overwrite_status = input('Already Exists. Overwrite?: ')
        if overwrite_status == 'rm':
            os.system('rm -rf ' + args.model_name)
        elif not 'y' in overwrite_status:
            exit(0)
    elif not os.path.exists(args.model_name):
        os.makedirs(args.model_name)

    return args

def get_args_infer_SIR(parser):
    parser.add_argument('--dataroot', type=str, default='./data/')
    parser.add_argument("--meta_root", type=str, default="./data/")
    parser.add_argument("--meta_file", type=str, default="DLRSD_retrieval_noise0.pkl")
    parser.add_argument('--dataset', type=str, choices=["DLRSD", "MultiScene-Clean", "ML-AID", "WHDLD", "PlanetUAS"],
                        default='DLRSD')
    parser.add_argument('--workers', type=int, default=32)
    parser.add_argument('--results_dir', type=str, default='results/')
    parser.add_argument("--noise_rate", type=float, default=0.0)
    parser.add_argument("--model_path", type=str, default="/home/jiangfanzhuo/SpaceIR/results/DLRSD.6layer.12heads.bsz_64.cls_eps_100.idx_eps_50.ret_eps_50.id_len_1.contr_l_0.01.distance_prc_5.class_loss_CE.adam3.5e-05.exp4_MLPdecoder_w_global")
    parser.add_argument("--codebook_path", type=str,
                        default="/home/jiangfanzhuo/SpaceIR/results/DLRSD.6layer.12heads.bsz_64.cls_eps_100.idx_eps_0.ret_eps_50.id_len_2.contr_l_0.1.adam3.5e-05.RQ_AR_update_encoder/codes.pth")
    parser.add_argument("--beam_size", type=int, default=5)
    parser.add_argument("--return_size", type=int, default=100)
    parser.add_argument("--test_batch_size", type=int, default=1)

    # Optimization
    parser.add_argument("--batch_size", type=int, default=64)
    parser.add_argument("--smoothing", type=float, default=0.1)

    # Model
    parser.add_argument('--layers', type=int, default=6)
    parser.add_argument('--heads', type=int, default=12)
    parser.add_argument('--dropout', type=float, default=0.1)
    parser.add_argument("--distance_prc", type=int, default=5)
    parser.add_argument('--id_len', type=int, default=1)
    parser.add_argument('--k_bit', type=int, default=8)
    parser.add_argument('--hidden_dim', type=int, default=768)
    parser.add_argument("--decoder_type", type=str, default="MLP")

    # Image Sizes
    parser.add_argument('--scale_size', type=int, default=640)
    parser.add_argument('--crop_size', type=int, default=576)

    parser.add_argument('--name', type=str, default='infer_model_exp1')
    args = parser.parse_args()

    model_name = args.dataset
    if args.dataset == 'DLRSD':
        args.num_labels = 17
    elif args.dataset == "MultiScene-Clean":
        args.num_labels = 36
    elif args.dataset == "ML-AID":
        args.num_labels = 17
    elif args.dataset == "WHDLD":
        args.num_labels = 6
    elif args.dataset == "PlanetUAS":
        args.num_labels = 14
    else:
        print('dataset not included')
        exit()

    if args.name != '':
        model_name += '.' + args.name

    model_name = os.path.join(args.results_dir, model_name)

    args.model_name = model_name


    return args

## EXP2
def get_args_train_SIR2(parser):
    parser.add_argument('--dataroot', type=str, default='./data/')
    parser.add_argument("--meta_root", type=str, default="./data/")
    parser.add_argument("--meta_file", type=str, default="DLRSD_retrieval_noise0.pkl")
    parser.add_argument('--dataset', type=str, choices=["DLRSD", "MultiScene-Clean", "ML-AID", "BigEarthNet-19", "WHDLD", "PlanetUAS"],
                        default='DLRSD')
    parser.add_argument('--workers', type=int, default=16)
    parser.add_argument('--results_dir', type=str, default='results/')
    parser.add_argument("--noise_rate", type=float, default=0.0)
    parser.add_argument("--cuda", type=str, default="1")

    # Optimization
    parser.add_argument('--optim', type=str, choices=['adam', 'sgd'], default='adam')
    parser.add_argument('--lr', type=float, default=3.5e-5)
    parser.add_argument("--batch_size", type=int, default=64)
    parser.add_argument("--classification_epochs", type=int, default=80)
    parser.add_argument("--retrieval_epochs", type=int, default=80)
    parser.add_argument("--smoothing", type=float, default=0.1)
    parser.add_argument("--contr_lambda", type=float, default=0.1)
    parser.add_argument("--contr_mu", type=float, default=0.1)
    parser.add_argument("--grad_ac_steps", type=int, default=1)
    parser.add_argument("--classification_resume", action="store_true")
    parser.add_argument("--classification_path", type=str, default="/home/jiangfanzhuo/SpaceIR/results/MultiScene-Clean.6layer.16heads.bsz_64.cls_eps_100.idx_eps_50.ret_eps_50.id_len_1.contr_l_0.1.contr_mu_0.1.distance_prc_10.class_loss_CE.adam3.5e-05.exp3_split_decoder/classification_model.pth")
    parser.add_argument("--indexing_resume", action="store_true")
    parser.add_argument("--indexing_path", type=str,
                        default="/home/jiangfanzhuo/SpaceIR/results/PlanetUAS.6layer.12heads.bsz_64.cls_eps_100.idx_eps_50.ret_eps_50.id_len_1.contr_l_0.1.distance_prc_50.class_loss_BCE.adam3.5e-05.exp1_split_decoder/indexing_model.pth")
    parser.add_argument("--class_loss", type=str, default="CE")
    parser.add_argument("--test_batch_size", type=int, default=128)

    # Model
    parser.add_argument('--layers', type=int, default=6)
    parser.add_argument('--heads', type=int, default=12)
    parser.add_argument('--dropout', type=float, default=0.1)
    parser.add_argument("--distance_prc", type=int, default=10)
    parser.add_argument('--hidden_dim', type=int, default=768)

    # Image Sizes
    parser.add_argument('--scale_size', type=int, default=640)
    parser.add_argument('--crop_size', type=int, default=576)

    # Testing Models
    parser.add_argument('--inference', action='store_true')
    parser.add_argument('--resume', action='store_true')
    parser.add_argument('--saved_model_name', type=str, default='')

    parser.add_argument('--overwrite', action='store_true')
    parser.add_argument('--name', type=str, default="SIR2_exp1")
    args = parser.parse_args()

    model_name = args.dataset
    if args.dataset == 'DLRSD':
        args.num_labels = 17
    elif args.dataset == "MultiScene-Clean":
        args.num_labels = 36
    elif args.dataset == "ML-AID":
        args.num_labels = 17
    elif args.dataset == "BigEarthNet-19":
        args.num_labels = 15
    elif args.dataset == "WHDLD":
        args.num_labels = 6
    elif args.dataset == "PlanetUAS":
        args.num_labels = 14
    else:
        print('dataset not included')
        exit()

    model_name += '.' + str(args.layers) + 'layer'
    model_name += '.' + str(args.heads) + 'heads'
    model_name += ".bsz_{}".format(int(args.batch_size))
    model_name += '.cls_eps_{}'.format(int(args.classification_epochs * args.grad_ac_steps))
    model_name += '.ret_eps_{}'.format(int(args.retrieval_epochs * args.grad_ac_steps))
    model_name += ".contr_l_{}".format(str(args.contr_lambda))
    model_name += ".contr_mu_{}".format(str(args.contr_mu))
    model_name += ".distance_prc_{}".format(int(args.distance_prc))
    model_name += ".class_loss_{}".format(str(args.class_loss))
    model_name += '.' + args.optim + str(args.lr)  # .split('.')[1]

    if args.name != '':
        model_name += '.' + args.name

    if not os.path.exists(args.results_dir):
        os.makedirs(args.results_dir)

    model_name = os.path.join(args.results_dir, model_name)

    args.model_name = model_name

    if args.inference:
        args.epochs = 1

    # global feature(hash)
    # find the value of ζ

    # workbook = openpyxl.load_workbook('codetable.xlsx')
    # sheet = workbook.active
    # args.threshold = sheet.cell(row=64 + 1, column=math.ceil(math.log(args.num_labels, 2)) + 1).value
    args.threshold = 0.5

    if os.path.exists(args.model_name) and (not args.overwrite) and (not 'test' in args.name) and (not eval) and (
    not args.inference) and (not args.resume):
        print(args.model_name)
        overwrite_status = input('Already Exists. Overwrite?: ')
        if overwrite_status == 'rm':
            os.system('rm -rf ' + args.model_name)
        elif not 'y' in overwrite_status:
            exit(0)
    elif not os.path.exists(args.model_name):
        os.makedirs(args.model_name)

    return args

def get_args_infer_SIR2(parser):
    parser.add_argument('--dataroot', type=str, default='./data/')
    parser.add_argument("--meta_root", type=str, default="./data/")
    parser.add_argument("--meta_file", type=str, default="DLRSD_retrieval_noise0.pkl")
    parser.add_argument('--dataset', type=str, choices=["DLRSD", "MultiScene-Clean", "ML-AID", "WHDLD", "PlanetUAS"],
                        default='DLRSD')
    parser.add_argument('--workers', type=int, default=32)
    parser.add_argument('--results_dir', type=str, default='results/')
    parser.add_argument("--noise_rate", type=float, default=0.0)
    parser.add_argument("--model_path", type=str, default="/home/jiangfanzhuo/SpaceIR/results/DLRSD.6layer.12heads.bsz_64.cls_eps_80.ret_eps_80.contr_l_0.1.contr_mu_0.1.distance_prc_10.class_loss_CE.adam3.5e-05.SIR2_exp1")
    parser.add_argument("--codebook_path", type=str,
                        default="/home/jiangfanzhuo/SpaceIR/results/DLRSD.6layer.12heads.bsz_64.cls_eps_100.idx_eps_0.ret_eps_50.id_len_2.contr_l_0.1.adam3.5e-05.RQ_AR_update_encoder/codes.pth")
    parser.add_argument("--beam_size", type=int, default=10)
    parser.add_argument("--return_size", type=int, default=100)
    parser.add_argument("--test_batch_size", type=int, default=1)

    # Optimization
    parser.add_argument("--batch_size", type=int, default=64)
    parser.add_argument("--smoothing", type=float, default=0.1)

    # Model
    parser.add_argument('--layers', type=int, default=6)
    parser.add_argument('--heads', type=int, default=12)
    parser.add_argument('--dropout', type=float, default=0.1)
    parser.add_argument("--distance_prc", type=int, default=10)
    parser.add_argument('--hidden_dim', type=int, default=768)

    # Image Sizes
    parser.add_argument('--scale_size', type=int, default=640)
    parser.add_argument('--crop_size', type=int, default=576)

    parser.add_argument('--name', type=str, default='SIR2_inference')
    args = parser.parse_args()

    model_name = args.dataset
    if args.dataset == 'DLRSD':
        args.num_labels = 17
    elif args.dataset == "MultiScene-Clean":
        args.num_labels = 36
    elif args.dataset == "ML-AID":
        args.num_labels = 17
    elif args.dataset == "WHDLD":
        args.num_labels = 6
    elif args.dataset == "PlanetUAS":
        args.num_labels = 14
    else:
        print('dataset not included')
        exit()

    if args.name != '':
        model_name += '.' + args.name

    model_name = os.path.join(args.results_dir, model_name)

    args.model_name = model_name


    return args

## Joint Learning
def get_args_train_SIR_joint(parser):
    parser.add_argument('--dataroot', type=str, default='./data/')
    parser.add_argument("--meta_root", type=str, default="./data/")
    parser.add_argument("--meta_file", type=str, default="DLRSD_retrieval_noise0.pkl")
    parser.add_argument('--dataset', type=str, choices=["DLRSD", "MultiScene-Clean", "ML-AID", "BigEarthNet-19", "WHDLD", "PlanetUAS", "MLRSNet"],
                        default='DLRSD')
    parser.add_argument('--workers', type=int, default=16)
    parser.add_argument('--results_dir', type=str, default='results/')
    parser.add_argument("--noise_rate", type=float, default=0)
    parser.add_argument("--cuda", type=str, default="0,1")

    # Optimization
    parser.add_argument('--optim', type=str, choices=['adam', 'sgd'], default='adam')
    parser.add_argument('--lr', type=float, default=3.5e-5)
    parser.add_argument("--batch_size", type=int, default=32)
    parser.add_argument("--epochs", type=int, default=100)
    parser.add_argument("--warm_epochs", type=int, default=50)
    parser.add_argument("--eval_step", type=int, default=1)
    parser.add_argument("--smoothing", type=float, default=0.1)
    parser.add_argument("--contr_lambda", type=float, default=0.1)
    # parser.add_argument("--contr_mu", type=float, default=0.1)
    parser.add_argument("--grad_ac_steps", type=int, default=1)
    parser.add_argument("--train_resume", action="store_true")
    parser.add_argument("--train_path", type=str, default="/home/jiangfanzhuo/SpaceIR/results/DLRSD.6layer.12heads.bsz_64.cls_eps_100.idx_eps_50.ret_eps_50.id_len_1.contr_l_0.1.contr_mu_0.1.distance_prc_10.class_loss_CE.adam3.5e-05.exp4_MLPdecoder_w_global/classification_model.pth")
    parser.add_argument("--warm_resume", action="store_true")
    parser.add_argument("--warm_path", type=str,
                        default="/home/jiangfanzhuo/SpaceIR/results/DLRSD.6layer.12heads.bsz_64.eps_100.warm_eps_50.id_len_1.contr_l_0.1.distance_prc_10.class_loss_CE.adam3.5e-05.exp4_jointV2_MLPdecoder_w_global/warm_model.pth")
    parser.add_argument("--class_loss", type=str, default="CE")
    parser.add_argument("--test_batch_size", type=int, default=128)

    # Model
    parser.add_argument('--layers', type=int, default=6)
    parser.add_argument('--heads', type=int, default=12)
    parser.add_argument('--dropout', type=float, default=0.1)
    parser.add_argument("--distance_prc", type=int, default=10)
    parser.add_argument('--id_len', type=int, default=1)
    parser.add_argument('--k_bit', type=int, default=8)
    parser.add_argument('--hidden_dim', type=int, default=768)
    parser.add_argument("--decoder_type", type=str, default="MLP")

    # Image Sizes
    parser.add_argument('--scale_size', type=int, default=640)
    parser.add_argument('--crop_size', type=int, default=576)

    # Testing Models
    parser.add_argument('--inference', action='store_true')
    parser.add_argument('--resume', action='store_true')
    parser.add_argument('--saved_model_name', type=str, default='')

    parser.add_argument('--overwrite', action='store_true')
    parser.add_argument('--name', type=str, default='exp4_jointV3_gumbel_MLPdecoder_w_global')
    args = parser.parse_args()

    model_name = args.dataset
    if args.dataset == 'DLRSD':
        args.num_labels = 17
    elif args.dataset == "MultiScene-Clean":
        args.num_labels = 36
    elif args.dataset == "ML-AID":
        args.num_labels = 17
    elif args.dataset == "BigEarthNet-19":
        args.num_labels = 15
    elif args.dataset == "WHDLD":
        args.num_labels = 6
    elif args.dataset == "PlanetUAS":
        args.num_labels = 14
    elif args.dataset == "MLRSNet":
        args.num_labels = 60
    else:
        print('dataset not included')
        exit()

    model_name += '.' + str(args.layers) + 'layer'
    model_name += '.' + str(args.heads) + 'heads'
    model_name += ".bsz_{}".format(int(args.batch_size))
    model_name += '.eps_{}'.format(int(args.epochs * args.grad_ac_steps))
    model_name += '.warm_eps_{}'.format(int(args.warm_epochs * args.grad_ac_steps))
    model_name += ".id_len_{}".format(int(args.id_len))
    model_name += ".contr_l_{}".format(str(args.contr_lambda))
    model_name += ".distance_prc_{}".format(int(args.distance_prc))
    model_name += ".class_loss_{}".format(str(args.class_loss))
    model_name += '.' + args.optim + str(args.lr)  # .split('.')[1]

    if args.name != '':
        model_name += '.' + args.name

    if not os.path.exists(args.results_dir):
        os.makedirs(args.results_dir)

    model_name = os.path.join(args.results_dir, model_name)

    args.model_name = model_name

    if args.inference:
        args.epochs = 1

    # global feature(hash)
    # find the value of ζ

    # workbook = openpyxl.load_workbook('codetable.xlsx')
    # sheet = workbook.active
    # args.threshold = sheet.cell(row=64 + 1, column=math.ceil(math.log(args.num_labels, 2)) + 1).value
    args.threshold = 0.5

    if os.path.exists(args.model_name) and (not args.overwrite) and (not 'test' in args.name) and (not eval) and (
    not args.inference) and (not args.resume):
        print(args.model_name)
        overwrite_status = input('Already Exists. Overwrite?: ')
        if overwrite_status == 'rm':
            os.system('rm -rf ' + args.model_name)
        elif not 'y' in overwrite_status:
            exit(0)
    elif not os.path.exists(args.model_name):
        os.makedirs(args.model_name)

    return args

def get_args_infer_SIR_joint(parser):
    parser.add_argument('--dataroot', type=str, default='./data/')
    parser.add_argument("--meta_root", type=str, default="./data/")
    parser.add_argument("--meta_file", type=str, default="DLRSD_retrieval_noise0.pkl")
    parser.add_argument('--dataset', type=str, choices=["DLRSD", "MultiScene-Clean", "ML-AID", "WHDLD", "PlanetUAS", "MLRSNet"],
                        default='DLRSD')
    parser.add_argument('--workers', type=int, default=32)
    parser.add_argument('--results_dir', type=str, default='results/')
    parser.add_argument("--noise_rate", type=float, default=0)
    parser.add_argument("--model_path", type=str, default="/home/jiangfanzhuo/SpaceIR/results/DLRSD.6layer.12heads.bsz_32.eps_100.warm_eps_50.id_len_1.contr_l_0.1.distance_prc_10.class_loss_CE.adam3.5e-05.exp4_jointV3_gumbel_MLPdecoder_w_global")
    parser.add_argument("--codebook_path", type=str,
                        default="/home/jiangfanzhuo/SpaceIR/results/DLRSD.6layer.12heads.bsz_64.cls_eps_100.idx_eps_0.ret_eps_50.id_len_2.contr_l_0.1.adam3.5e-05.RQ_AR_update_encoder/codes.pth")
    parser.add_argument("--beam_size", type=int, default=10)
    parser.add_argument("--return_size", type=int, default=100)
    parser.add_argument("--test_batch_size", type=int, default=1)

    # Optimization
    parser.add_argument("--batch_size", type=int, default=64)
    parser.add_argument("--smoothing", type=float, default=0.1)

    # Model
    parser.add_argument('--layers', type=int, default=6)
    parser.add_argument('--heads', type=int, default=12)
    parser.add_argument('--dropout', type=float, default=0.1)
    parser.add_argument("--distance_prc", type=int, default=10)
    parser.add_argument('--id_len', type=int, default=1)
    parser.add_argument('--k_bit', type=int, default=8)
    parser.add_argument('--hidden_dim', type=int, default=768)
    parser.add_argument("--decoder_type", type=str, default="MLP")

    # Image Sizes
    parser.add_argument('--scale_size', type=int, default=640)
    parser.add_argument('--crop_size', type=int, default=576)

    parser.add_argument('--name', type=str, default='infer_model_joint')
    args = parser.parse_args()

    model_name = args.dataset
    if args.dataset == 'DLRSD':
        args.num_labels = 17
    elif args.dataset == "MultiScene-Clean":
        args.num_labels = 36
    elif args.dataset == "ML-AID":
        args.num_labels = 17
    elif args.dataset == "WHDLD":
        args.num_labels = 6
    elif args.dataset == "PlanetUAS":
        args.num_labels = 14
    elif args.dataset == "MLRSNet":
        args.num_labels = 60
    else:
        print('dataset not included')
        exit()

    if args.name != '':
        model_name += '.' + args.name

    model_name = os.path.join(args.results_dir, model_name)

    args.model_name = model_name


    return args

## Cross Modal Learning
def get_args_train_SIRMM(parser):
    parser.add_argument('--dataroot', type=str, default='./data/')
    parser.add_argument("--meta_root", type=str, default="./data/")
    parser.add_argument("--meta_file", type=str, default="DOTA_retrieval_noise0.pkl")
    parser.add_argument('--dataset', type=str, default='MultiScene-Clean')
    parser.add_argument('--workers', type=int, default=16)
    parser.add_argument('--results_dir', type=str, default='results/')
    parser.add_argument("--noise_rate", type=float, default=0)

    # Optimization
    parser.add_argument('--optim', type=str, choices=['adam', 'sgd'], default='adam')
    parser.add_argument('--lr', type=float, default=3.5e-5)
    parser.add_argument("--batch_size", type=int, default=32) # single gpu batch size
    parser.add_argument("--epochs", type=int, default=30)
    parser.add_argument("--warm_epochs", type=int, default=50)
    parser.add_argument("--eval_step", type=int, default=5)
    parser.add_argument("--smoothing", type=float, default=0.1)
    parser.add_argument("--contr_lambda", type=float, default=0.1)
    parser.add_argument("--pretrain_path", type=str, default="None")
    parser.add_argument("--train_resume", action="store_true")
    parser.add_argument("--train_path", type=str, default="/home/jiangfanzhuo/SpaceIR/results/DLRSD.6layer.12heads.bsz_32.eps_100.warm_eps_50.hidden_768.contr_l_0.1.distance_prc_10.class_loss_BCE.adam3.5e-05.MM_test/retrieval_model.pth")
    parser.add_argument("--warm_resume", action="store_true")
    parser.add_argument("--warm_path", type=str,
                        default="/home/jiangfanzhuo/SpaceIR/results/DLRSD.6layer.12heads.bsz_32.eps_100.warm_eps_50.hidden_768.contr_l_0.1.distance_prc_10.class_loss_BCE.adam3.5e-05.MM_test/warm_model.pth")
    parser.add_argument("--class_loss", type=str, default="BCE")
    parser.add_argument("--test_batch_size", type=int, default=1)
    parser.add_argument("--max_topK", type=int, default=10)

    parser.add_argument("--test_noise", action="store_true")

    # Model
    parser.add_argument('--image_token', type=int, default=324)
    parser.add_argument('--text_token', type=int, default=150)
    parser.add_argument('--layers', type=int, default=6)
    parser.add_argument('--heads', type=int, default=12)
    parser.add_argument('--dropout', type=float, default=0.1)
    parser.add_argument("--distance_prc", type=int, default=10)
    parser.add_argument('--hidden_dim', type=int, default=768)

    parser.add_argument("--text_encoder_model_path", type=str, default="/mnt/data/jiangfanzhuo/Weights_file/bert-base-uncased")
    parser.add_argument("--image_encoder_model", type=str,
                        default="resnet50")
    parser.add_argument('--decoder_forward_dim', type=int, default=3072)

    # Ablation exp
    parser.add_argument("--abl_setting", type=str, default="none")

    # Image Sizes
    parser.add_argument('--scale_size', type=int, default=640)
    parser.add_argument('--crop_size', type=int, default=576)

    parser.add_argument('--name', type=str, default='MM_ddp')
    args = parser.parse_args()

    model_name = args.dataset
    if args.dataset == 'DLRSD':
        args.num_labels = 17
    elif args.dataset == "MultiScene-Clean":
        args.num_labels = 36
    elif args.dataset == "ML-AID":
        args.num_labels = 17
    elif args.dataset == "BigEarthNet-19":
        args.num_labels = 15
    elif args.dataset == "WHDLD":
        args.num_labels = 6
    elif args.dataset == "PlanetUAS":
        args.num_labels = 14
    elif args.dataset == "MLRSNet":
        args.num_labels = 60
    elif args.dataset == "DOTA":
        args.num_labels = 16
    elif args.dataset == "RICE-CG":
        args.num_labels = 8
    elif args.dataset == "RICE-SIR":
        args.num_labels = 8
    elif args.dataset == "Cloud-SIR":
        args.num_labels = 17
    elif args.dataset == "CUHK-SIR":
        args.num_labels = 13
    else:
        print('dataset not included')
        exit()

    model_name += '.' + str(args.layers) + 'layer'
    model_name += '.' + str(args.heads) + 'heads'
    model_name += ".bsz_{}".format(int(args.batch_size))
    model_name += '.eps_{}'.format(int(args.epochs))
    model_name += '.warm_eps_{}'.format(int(args.warm_epochs))
    model_name += ".hidden_{}".format(int(args.hidden_dim))
    model_name += ".contr_l_{}".format(str(args.contr_lambda))
    model_name += ".distance_prc_{}".format(int(args.distance_prc))
    model_name += ".class_loss_{}".format(str(args.class_loss))
    model_name += '.' + args.optim + str(args.lr)  # .split('.')[1]
    model_name += ".abl_set_{}".format(args.abl_setting)

    if args.name != '':
        model_name += '.' + args.name

    if not os.path.exists(args.results_dir):
        os.makedirs(args.results_dir)

    model_name = os.path.join(args.results_dir, model_name)

    args.model_name = model_name
    os.makedirs(args.model_name, exist_ok=True)
    args.threshold = 0.5

    return args

def get_args_infer_SIRMM(parser):
    parser.add_argument('--dataroot', type=str, default='./data/')
    parser.add_argument("--meta_root", type=str, default="./data/")
    parser.add_argument("--meta_file", type=str, default="MultiScene-Clean_retrieval_noise0.pkl")
    parser.add_argument('--dataset', type=str, default='Cloud-SIR')
    parser.add_argument('--workers', type=int, default=16)
    parser.add_argument('--results_dir', type=str, default='results/')
    parser.add_argument("--model_path", type=str,
                            default="/home/jiangfanzhuo/SpaceIR/results/MultiScene-Clean.6layer.12heads.bsz_32.eps_30.warm_eps_50.hidden_768.contr_l_0.1.distance_prc_10.class_loss_BCE.adam3.5e-05.MM_test")
    parser.add_argument("--beam_size", type=int, default=10)
    parser.add_argument("--return_size", type=int, default=100)
    parser.add_argument("--max_topK", type=int, default=100)
    parser.add_argument("--test_batch_size", type=int, default=1)
    parser.add_argument("--test_noise", action="store_true")
    parser.add_argument("--noise_rate", type=int, default=1)

    # Optimization
    parser.add_argument("--batch_size", type=int, default=32)
    parser.add_argument("--smoothing", type=float, default=0.1)

    # Model
    parser.add_argument('--image_token', type=int, default=324)
    parser.add_argument('--text_token', type=int, default=150)
    parser.add_argument('--layers', type=int, default=6)
    parser.add_argument('--heads', type=int, default=12)
    parser.add_argument('--dropout', type=float, default=0.1)
    parser.add_argument("--distance_prc", type=int, default=10)
    parser.add_argument('--hidden_dim', type=int, default=768)

    # Ablation exp
    parser.add_argument("--abl_setting", type=str, default="none")

    parser.add_argument("--text_encoder_model_path", type=str, default="/mnt/data/jiangfanzhuo/Weights_file/bert-base-uncased")
    parser.add_argument("--image_encoder_model", type=str,
                        default="resnet50")
    parser.add_argument('--decoder_forward_dim', type=int, default=3072)

    # Image Sizes
    parser.add_argument('--scale_size', type=int, default=640)
    parser.add_argument('--crop_size', type=int, default=576)

    args = parser.parse_args()

    model_name = str(args.model_path).split('/')[-1]
    if args.dataset == 'DLRSD':
        args.num_labels = 17
    elif args.dataset == "MultiScene-Clean":
        args.num_labels = 36
    elif args.dataset == "ML-AID":
        args.num_labels = 17
    elif args.dataset == "BigEarthNet-19":
        args.num_labels = 15
    elif args.dataset == "WHDLD":
        args.num_labels = 6
    elif args.dataset == "PlanetUAS":
        args.num_labels = 14
    elif args.dataset == "MLRSNet":
        args.num_labels = 60
    elif args.dataset == "DOTA":
        args.num_labels = 16
    elif args.dataset == "RICE-CG":
        args.num_labels = 8
    elif args.dataset == "RICE-SIR":
        args.num_labels = 8
    elif args.dataset == "Cloud-SIR":
        args.num_labels = 17
    else:
        print('dataset not included')
        exit()

    if not os.path.exists(args.results_dir):
        os.makedirs(args.results_dir)

    model_name = os.path.join(args.results_dir, model_name, "inference")

    args.model_name = model_name
    os.makedirs(args.model_name, exist_ok=True)
    args.threshold = 0.5

    return args